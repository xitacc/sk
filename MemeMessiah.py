import requests
import sys
import time
from datetime import datetime

# t0 = time.time()*1000
# # r = requests.get('http://194.95.50.48/values/')
# r = requests.get('https://boards.4chan.org/pol/thread/278219668')
#
# # r = requests.get('http://127.0.0.1:80/values/')
# t1 = time.time()*1000
# print("req Time="+str(t1-t0))
# print(len(r.text.encode('utf-8')))
# # print(r.text)
# print(r.content)
# //*[@id="t278219668"]
# html body.is_thread.board_pol.tomorrow.nws form#delform div.board div#t278219668.thread

# original_stdout = sys.stdout # Save a reference to the original standard output
# with open('valuesOutput80022.txt', 'w') as f:
#     sys.stdout = f # Change the standard output to the file we created.
#     print(r.text)
#     sys.stdout = original_stdout # Reset the standard output to its original value
from openpyxl import load_workbook
wb = load_workbook("20200120_Regionaldatenbank 1.2NAKorrigiertV2.xlsx", data_only=True)

# try:
#     sheet = wb['Lausitzdatenbank_mn_cp']
#     print(sheet)
# except KeyError:
#     print("Leider hat die Exceldatei keine Tabelle mit dem Namen Lausitzdatenbank_mn_cp ")
# sheet = wb['Lausitzdatenbank_mn_cpp']

sheet = wb['Lausitzdatenbank_mn_cp']
print(sheet)
print(neuerName)
print("Ja es geht weiter")
a=3+3

#
# from datetime import datetime
#
# now = datetime.now()
#
# wb.save("xDDDDDDDDDDDDD.xlsx")
#
# rIdx = 6
# row = next(sheet.iter_rows(min_row=rIdx, max_row=rIdx, min_col=1, max_col=149, values_only=True))
# print(row)
#
# rIdx = 7
# row = next(sheet.iter_rows(min_row=rIdx, max_row=rIdx, min_col=1, max_col=149, values_only=True))
# print(row)
#
# rIdx = 8
# row = next(sheet.iter_rows(min_row=rIdx, max_row=rIdx, min_col=1, max_col=149, values_only=True))
# print(row)
#
